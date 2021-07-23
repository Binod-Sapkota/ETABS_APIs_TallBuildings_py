# This file contains areaAll object class and Story Object Class
import numpy as np
import os
import pandas as pd

class ObjEtabsFile:
    def __init__(self, directoryName):
        self.files = []
        self.names = []
        self.dirs = []
        self.roots = []
        for root, dira, file in os.walk(directoryName):
            for filename in file:
                if filename.endswith(".EDB"):
                    self.names.append(filename)
                    self.roots.append(root)
                    self.files.append(os.path.join(root, filename))


# ======================================================================================================================

from openpyxl import load_workbook


# sheetname = "model"
# filename = "model.xlsx"
def getframeData_excel(filename, sheetname, totalElements):
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]

    # FrameName = []
    XI = []
    YI = []
    ZI = []
    XJ = []
    YJ = []
    ZJ = []
    PropName = []
    # UserName = []

    for i, row in enumerate(ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=totalElements + 2)):
        # FrameName.append(row[0].value)
        XI.append(row[0].value)
        YI.append(row[1].value)
        ZI.append(row[2].value)
        XJ.append(row[3].value)
        YJ.append(row[4].value)
        ZJ.append(row[5].value)
        PropName.append(row[6].value)
        # UserName.append(row[8].value)

    add_frame_dict = {"XI": XI,
                      "YI": YI,
                      "ZI": ZI,
                      "XJ": XJ,
                      "YJ": YJ,
                      "ZJ": ZJ,
                      "PropName": PropName}  # "FrameName": FrameName,,"UserName": UserName}
    return add_frame_dict


# ======================================================================================================================

def getmaterialdata_excel(filename, sheetname):
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]

    MatType = []
    Region = []
    Standard = []
    Grade = []
    UserName = []

    for i, row in enumerate(ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=ws.max_row)):
        MatType.append(row[0].value)
        Region.append(row[1].value)
        Standard.append(row[2].value)
        Grade.append(row[3].value)
        UserName.append(row[4].value)

    material_dict = {"MatType": MatType,
                     "Region": Region,
                     "Standard": Standard,
                     "Grade": Grade,
                     "UserName": UserName}
    return material_dict


# ======================================================================================================================

def getsectiondata_excel(filename, sheetname):
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]

    Name = []
    MatProp = []
    T3 = []
    T2 = []

    for i, row in enumerate(ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=ws.max_row)):
        Name.append(row[0].value)
        MatProp.append(row[1].value)
        T3.append(row[2].value)
        T2.append(row[3].value)

    section_dict = {"Name": Name,
                    "MatProp": MatProp,
                    "T3": T3,
                    "T2": T2}
    return section_dict


def getloaddata_excel(filename, sheetname):
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]

    loadName = []
    loadPat = []
    loadValue = []

    for i, row in enumerate(ws.iter_rows(min_row=3, max_col=8, max_row=12)):
        loadName.append(row[0].value)
        loadPat.append(row[1].value)
        loadValue.append([ro.value for ro in row[2:8]])

    load_dict = {"loadName": loadName,
                 "loadPat": loadPat,
                 "loadValue": loadValue}
    return load_dict


# ======================================================================================================================


def getModelSetting(filename="model.xlsx", sheetname="setting", numberOfModels=None):
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]
    modelNumber = []
    originX = []
    originY = []
    originZ = []
    numberStory = []
    typicalStoryH = []
    bottomStoryH = []
    lineX = []
    lineY = []
    spacingX = []
    spacingY = []

    colName = []
    colMatProp = []
    coldepth_x = []
    colwidth_y = []

    beamName = []
    beamMatProp = []
    beamdepth = []
    beamwidth = []

    if numberOfModels is None:
        rowMax = ws.max_row
    else:
        rowMax = numberOfModels + 3  # 4 is for starting of row

    for row in ws.iter_rows(min_row=4, max_col=ws.max_column, max_row=rowMax):
        modelNumber.append(row[0].value)
        originX.append(row[1].value)
        originY.append(row[2].value)
        originZ.append(row[3].value)
        numberStory.append(row[4].value)
        typicalStoryH.append(row[5].value)
        bottomStoryH.append(row[6].value)
        lineX.append(row[7].value)
        lineY.append(row[8].value)
        spacingX.append(row[9].value)
        spacingY.append(row[10].value)
        colName.append(row[11].value)
        colMatProp.append(row[12].value)
        coldepth_x.append(row[13].value)
        colwidth_y.append(row[14].value)
        beamName.append(row[15].value)
        beamMatProp.append(row[16].value)
        beamdepth.append(row[17].value)
        beamwidth.append(row[18].value)

    # origin_dict = {"modelNumber": modelNumber,
    #                "originX": originX,
    #                "originY": originY,
    #                "originZ": originZ}
    #
    # gridInfo_dict = {"numberStory": numberStory,
    #                  "typicalStoryH": typicalStoryH,
    #                  "bottomStoryH": bottomStoryH,
    #                  "lineX": lineX,
    #                  "lineY": lineY,
    #                  "spacingX": spacingX,
    #                  "spacingY": spacingY}

    modelInfo = np.transpose(np.array([modelNumber, originX, originY, originZ, numberStory, typicalStoryH,
                                       bottomStoryH, lineX, lineY, spacingX, spacingY]))

    sectionInfo = np.transpose(np.array([colName, colMatProp, coldepth_x, colwidth_y,
                                         beamName, beamMatProp, beamdepth, beamwidth]))
    return modelInfo, sectionInfo


# ======================================================================================================================


def getPointCoordinate(modelInfo):
    # originX = origin_dict['originX'][0]
    # originY = origin_dict['originY'][0]
    # originZ = origin_dict['originZ'][0]
    #
    # numberStory = int(gridInfo_dict['numberStory'][0])
    # typicalStoryH = float(gridInfo_dict['typicalStoryH'][0])
    # bottomStoryH = float(gridInfo_dict['bottomStoryH'][0])
    # lineX = int(gridInfo_dict['lineX'][0])
    # lineY = int(gridInfo_dict['lineY'][0])
    # spacingX = float(gridInfo_dict['spacingX'][0])
    # spacingY = float(gridInfo_dict['spacingY'][0])

    modelNumber, originX, originY, originZ, numberStory, \
    typicalStoryH, bottomStoryH, lineX, lineY, spacingX, spacingY = modelInfo

    numberStory = int(numberStory)
    lineX = int(lineX)
    lineY = int(lineY)

    totalpoints = lineX * lineY * (numberStory + 1)

    xCoordinate = np.zeros(totalpoints)
    yCoordinate = np.zeros(totalpoints)
    zCoordinate = np.zeros(totalpoints)
    # yCoordinate = np.zeros((totalpoints, 1))
    # zCoordinate = np.zeros((totalpoints, 1))
    count = 0
    startIndex = []
    endIndex = []
    for i in range(lineX):
        for j in range(lineY):
            for k in range(numberStory + 1):
                xCoordinate[count] = originX + spacingX * i
                yCoordinate[count] = originY + spacingY * j
                zCoordinate[count] = originZ + typicalStoryH * k

                count = count + 1

    jointCoordinate = np.transpose(np.array([xCoordinate, yCoordinate, zCoordinate]))

    jointCoordinate = np.array(jointCoordinate)
    leftJoints = []
    for i in range(np.shape(jointCoordinate)[0]):
        if jointCoordinate[i, 0] == originX and jointCoordinate[i, 1] == originZ and jointCoordinate[i, 2] != 0:
            leftJoints.append(i + 1)  # list(jointCoordinate[i,:]))

    # leftJoints are where lateral load is applied
    leftJoints = np.transpose(np.array(leftJoints, ndmin=2))

    # topRightCornerJoint is used to measure drift,and displacement
    topRightCornerJoint = str(int(np.size(jointCoordinate) / 3))

    # dfCoordinates = pd.DataFrame(jointCoordinate, columns=["x", "y", 'z'])

    # dfCoordinates.to_csv("Coordinate.csv")

    # ==================Plot coordinates along with numbering=========================================
    # pltbool = False
    # if pltbool:
    #     import matplotlib.pyplot as plt
    #     from mpl_toolkits.mplot3d import Axes3D
    #
    #     fig = plt.figure()
    #     ax = fig.add_subplot(111, projection='3d')
    #     ax.scatter(xCoordinate, yCoordinate, zCoordinate)
    #     for zdir, x, y, z in zip(range(len(xCoordinate)), xCoordinate, yCoordinate, zCoordinate):
    #         label = zdir
    #         ax.text(x, y, z, label, None)
    #     # ax.text(xCoordinate,yCoordinate,zCoordinate, str(range(len(xCoordinate))))
    #     plt.show()

    return jointCoordinate, topRightCornerJoint, leftJoints


# ======================================================================================================================


def get_ElementCoordinate(jointCoordinate):  # jointCoordinate should be a numpy array with size (# points,3)
    startCO = []
    endCO = []
    secProp = []
    frameName = []
    plot_option = False

    df = pd.DataFrame(jointCoordinate, columns=["x", "y", "z"])

    # get only the unique values of x, y and z coordinate
    x_unique = pd.unique(df['x'])
    y_unique = pd.unique(df['y'])
    z_unique = pd.unique(df['z'])

    # for each unique value of x, and then for each unique value of y the coordinate is
    # obtained which represent the coordinate in one plane

    for x in x_unique:
        dfx = df[df['x'] == x]
        for y in y_unique:
            dfy = dfx[dfx["y"] == y]

            # get starting and ending point for a frame from a points in one plane
            for i in range(len(dfy) - 1):
                startCO.append(pd.Series.to_list(dfy.iloc[i]))
                endCO.append(pd.Series.to_list(dfy.iloc[i + 1]))
                secProp.append("COL")

    # plt.show()
    for y in y_unique:
        dfy = df[df['y'] == y]

        for z in z_unique:
            if z != 0:
                dfz = dfy[dfy["z"] == z]

                for i in range(len(dfz) - 1):
                    startCO.append(pd.Series.to_list(dfz.iloc[i]))
                    endCO.append(pd.Series.to_list(dfz.iloc[i + 1]))
                    secProp.append("BEAM")

    for x in x_unique:
        dfx = df[df['x'] == x]
        for z in z_unique:
            if z != 0:
                dfz = dfx[dfx["z"] == z]

                for i in range(len(dfz) - 1):
                    startCO.append(pd.Series.to_list(dfz.iloc[i]))
                    endCO.append(pd.Series.to_list(dfz.iloc[i + 1]))
                    secProp.append("BEAM")

    elmCO = np.concatenate((np.array(startCO), np.array(endCO)), axis=1)
    elmCO_df = pd.DataFrame(elmCO, columns=['XI', 'YI', 'ZI', 'XJ', 'YJ', 'ZJ'])
    elmCO_df['PropName'] = secProp

    elmCO = elmCO_df.to_numpy()

    if plot_option:

        import matplotlib.pyplot as plt
        # from mpl_toolkits.mplot3d import Axes3D
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')

        for i in range(len(elmCO_df.index)):
            ax.plot((elmCO_df['XI'].iloc[i], elmCO_df['XJ'].iloc[i]), (elmCO_df['YI'].iloc[i], elmCO_df['YJ'].iloc[i]),
                    (elmCO_df['ZI'].iloc[i], elmCO_df['ZJ'].iloc[i]))

        plt.axis('off')
        plt.show()

    # elmCO_df.to_csv("elementcoordinate.csv")

    return elmCO


# ======================================================================================================================


def generateFrameData(modelInfo, filename, sheetname):
    jointCoordinate, topRightCornerJoint, leftJoints = getPointCoordinate(modelInfo)
    elmCO = get_ElementCoordinate(jointCoordinate)
    totalElements = np.shape(elmCO)[0]
    writedata_excel(leftJoints, filename, "load", startrange=(2, 0))
    writedata_excel(elmCO, filename, sheetname, startrange=(2, 0))
    return topRightCornerJoint, totalElements


# ======================================================================================================================


def drawNewModel(SapModel, ModelPath, filename, modelInfo, sectionInfo):
    # STEP1: initialize model and specify unit
    SapModel.InitializeNewModel(Units=6)

    # # create new blank model
    # ret = SapModel.File.NewBlank()

    # kN_m_C = 6
    # ret = SapModel.SetPresentUnits(Units=6)

    # get model setting
    # model origin

    # origin_dict, gridInfo_dict = getModelSetting(filename="model.xlsx", sheetname="setting")
    # originIndex = origin_dict['modelNumber'].index(modelnumber)
    # originX = origin_dict['originX'][originIndex]
    # originY = origin_dict['originY'][originIndex]
    # originZ = origin_dict['originZ'][originIndex]

    # model story info
    modelNumber, originX, originY, originZ, numberStory, \
    typicalStoryH, bottomStoryH, lineX, lineY, spacingX, spacingY = modelInfo

    # for i in range(1):
    #     numberStory = gridInfo_dict['numberStory'][i]
    #     typicalStoryH = gridInfo_dict['typicalStoryH'][i]
    #     bottomStoryH = gridInfo_dict['bottomStoryH'][i]
    #     lineX = gridInfo_dict['lineX'][i]
    #     lineY = gridInfo_dict['lineY'][i]
    #     spacingX = gridInfo_dict['spacingX'][i]
    #     spacingY = gridInfo_dict['spacingY'][i]

    ret = SapModel.SetPresentUnits_2(4, 6, 2)

    # STEP2 : MAKE GRIDS

    ret = SapModel.File.NewGridOnly(NumberStorys=int(numberStory), TypicalStoryHeight=typicalStoryH,
                                    BottomStoryHeight=bottomStoryH, NumberLinesX=int(lineX),
                                    NumberLinesY=int(lineY), SpacingX=spacingX, SpacingY=spacingY)

    # Set units
    # kN_m_C = 6
    # ret = SapModel.SetPresentUnits(6)

    # STEP 3: define material property
    material_dict = getmaterialdata_excel(filename, 'material')
    ret = SapModel.PropMaterial.AddMaterial(MatType=material_dict["MatType"][0], Region=material_dict["Region"][0],
                                            Standard=material_dict["Standard"][0], Grade=material_dict["Grade"][0],
                                            UserName=material_dict["UserName"][0])

    # STEP 4: define rectangular frame section property
    for i in range(2):
        secName = sectionInfo[0 + i * 4]
        secMatProp = sectionInfo[1 + i * 4]
        secT3 = float(sectionInfo[2 + i * 4])
        secT2 = float(sectionInfo[3 + i * 4])
        ret = SapModel.PropFrame.SetRectangle(Name=secName, MatProp=secMatProp, T3=secT3, T2=secT2)

    # STEP 5: add load patterns
    LTYPE_OTHER = 8
    LTYPE_DEAD = 1
    LTYPE_QUAKE = 5

    ret = SapModel.LoadPatterns.Add('SELF WEIGHT', LTYPE_DEAD, 1, True)
    ret = SapModel.LoadPatterns.Add('LATERAL', LTYPE_OTHER, 0, True)
    ret = SapModel.LoadPatterns.Add('EQ', LTYPE_QUAKE, 0, True)

    # STEP 6: DEFINE AUTOSIESMIC LOAD TO USER COEFFICIENT
    defineAutoSeismic_UserCoefficient(SapModel)

    # Auto-generate frame Data
    topRightCornerJoint, totalElements = generateFrameData(modelInfo, filename, 'model')

    # STEP 7: add frame object by coordinates
    framedata = getframeData_excel(filename, 'model', totalElements)

    for i in range(len(framedata['XI'])):
        # frmName = framedata['FrameName'][i]
        XI = framedata['XI'][i] + originX
        YI = framedata['YI'][i] + originY
        ZI = framedata['ZI'][i] + originZ
        XJ = framedata['XJ'][i] + originX
        YJ = framedata['YJ'][i] + originY
        ZJ = framedata['ZJ'][i] + originZ
        PropName = str(framedata['PropName'][i])
        # UserName = str(framedata['UserName'][i])

        [frmName, ret] = SapModel.FrameObj.AddByCoord(XI, YI, ZI, XJ, YJ, ZJ, PropName=PropName)

        # STEP 8: assign point object restraint at base
        Restraint = [True, True, True, True, True, True]

        if ZI == 0:
            [PointName1, PointName2, ret] = SapModel.FrameObj.GetPoints(frmName)
            ret = SapModel.PointObj.SetRestraint(PointName1, Restraint)

    # STEP 9: Define Diaphragm
    set_diaphragm(SapModel)

    # STEP 10: DEFINE MASS SOURCE
    setMassSource(SapModel)

    # STEP 11: SET MODEL LOAD CASE
    ret = setModalCase(SapModel)
    # set bucking load case
    setbucklingLoadCase(SapModel)

    # #Set Stories
    # inStoryNames = ["MyStory1", "MyStory2", "MyStory3", "MyStory4", "MyStory5", "MyStory6",
    #                 "MyStory7", "MyStory8", "MyStory9", "MyStory10"]
    # inStoryHeights = [3,3, 3, 3, 3, 3,3,3,3,3]
    #
    # ret = SapModel.Story.SetStories_2(BaseElevation=3.0, NumberStories=10,
    #                                   StoryNames=inStoryNames, StoryHeights=inStoryHeights)

    # refresh view, update (initialize) zoom
    ret = SapModel.View.RefreshView(0, False)

    # STEP 12: assign loading
    load_dict = getloaddata_excel(filename, 'load')
    for i in range(len(load_dict['loadName'])):
        loadName = str(load_dict['loadName'][i])
        loadPat = load_dict['loadPat'][i]
        loadValue = load_dict['loadValue'][i]
        ret = SapModel.PointObj.SetLoadForce(Name=loadName, LoadPat=loadPat, Value=loadValue)

    # STEP 13: save model
    ret = SapModel.File.Save(ModelPath)
    return topRightCornerJoint


# ======================================================================================================================

def analysisSetup(SapModel):
    # get load case run flags
    [NumberItems, CaseName, Run, ret] = SapModel.Analyze.GetRunCaseFlag()

    # set loadcase to run
    for loadcase in CaseName:
        if loadcase == "LATERAL" or loadcase == "EQ" or loadcase == "SELF WEIGHT" or \
                loadcase == "Modal" or loadcase == "buckling":
            ret = SapModel.Analyze.SetRunCaseFlag(loadcase, True)
        else:
            ret = SapModel.Analyze.SetRunCaseFlag(loadcase, False)

    # set active DOF (the below active DOF is for plane frame structure only)
    DOF = [True, False, True, False, True, False]
    ret = SapModel.Analyze.SetActiveDOF(DOF)

    ret = SapModel.Analyze.RunAnalysis()
    ret = SapModel.View.RefreshView(0, False)

    # ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    # ret = SapModel.Results.Setup.SetCaseSelectedForOutput("LATERAL")


# ======================================================================================================================


def writedata_excel(resultData, filename, sheetname, startrange=(0, 0)):
    # resultData should be numpy array

    wb = load_workbook(filename=filename, read_only=False)
    ws = wb[sheetname]
    s_row, s_column = startrange
    for row in range(np.size(resultData, 0)):
        for col in range(np.size(resultData, 1)):
            ws.cell(column=col + 1 + s_column, row=row + s_row + 1, value=resultData[row, col])

    wb.save(filename)


# ======================================================================================================================


def writedata_excel2(resultData, filename, sheetname, startrange=(0, 0), to_clear=False):
    # resultData should be numpy array
    import xlwings as xw

    os.path.exists(filename)
    if os.path.exists(filename):
        wb = xw.Book(filename)
    else:
        wb = xw.Book()
    sh = wb.sheets
    if sheetname in [sh[i].name for i in range(sh.count)]:
        ws = wb.sheets[sheetname]
    else:
        ws = wb.sheets.add(sheetname)

    if to_clear:
        ws.cells.clear_contents()
    ws.range(startrange[0], startrange[1]).value = resultData
    wb.save(filename)

    return (wb.app, wb)


# ======================================================================================================================


def defineAutoSeismic_UserCoefficient(SapModel):
    # TODO .. get_tableData automatically
    # LTYPE_QUAKE = 5
    # ret = SapModel.LoadPatterns.Add('EQ', LTYPE_QUAKE, 0, True)
    TableVersion = 1
    NumberRecords = 1
    tTableKey = "Load Pattern Definitions - Auto Seismic - User Coefficient"
    fieldKeysIncluded = ['Name', 'Is Auto Load', 'X Dir?', 'X Dir Plus Ecc?', 'X Dir Minus Ecc?', 'Y Dir?',
                         'Y Dir Plus Ecc?', 'Y Dir Minus Ecc?', 'Ecc Ratio', 'Top Story', 'Bottom Story', 'C', 'K']

    TableData = ['EQ', 'No', 'Yes', "no", 'No', 'No', 'No', 'No', '0.0', 'Story10', 'Base', '0.3', '2']

    [TableVersion, FieldsKeysIncluded, TableData,
     ret] = SapModel.DatabaseTables.SetTableForEditingArray(tTableKey, TableVersion,
                                                            fieldKeysIncluded, NumberRecords, TableData)
    [NumFatalErrors, NumErrorMsgs, NumWarnMsgs, NumInfoMsgs,
     ImportLog, ret] = SapModel.DatabaseTables.ApplyEditedTables(FillImportLog=True)


# ======================================================================================================================


def setModalCase(SapModel, nMode=30):
    # TODO .. get_tableData automatically
    # LTYPE_QUAKE = 5
    # ret = SapModel.LoadPatterns.Add('EQ', LTYPE_QUAKE, 0, True)

    ret = SapModel.LoadCases.ModalEigen.SetCase("Modal")

    TableVersion = 1
    NumberRecords = 1
    tTableKey = "Modal Case Definitions - Eigen"
    fieldKeysIncluded = ['Name', 'Max Modes']

    TableData = ['Modal', str(nMode)]

    [TableVersion, FieldsKeysIncluded, TableData,
     ret] = SapModel.DatabaseTables.SetTableForEditingArray(tTableKey, TableVersion,
                                                            fieldKeysIncluded, NumberRecords, TableData)

    [NumFatalErrors, NumErrorMsgs, NumWarnMsgs, NumInfoMsgs,
     ImportLog, ret] = SapModel.DatabaseTables.ApplyEditedTables(FillImportLog=True)
    return ret


# ======================================================================================================================


def set_diaphragm(SapModel):
    [StoryNumber, StoryName, ret] = SapModel.Story.GetNameList()
    LenX = []
    LenY = []
    for story in StoryName:
        [NumberNames, pointNameinStorylist, ret] = SapModel.PointObj.GetNameListOnStory(story)
        X = []
        Y = []
        Z = []

        for pointName in pointNameinStorylist:
            ret = SapModel.PointObj.SetDiaphragm(Name=pointName, DiaphragmOption=3, DiaphragmName="D1")


# ======================================================================================================================


def setMassSource(SapModel):
    # TODO .. get_tableData automatically

    TableVersion = 1
    NumberRecords = 1
    tTableKey = "Mass Source Definition"
    fieldKeysIncluded = ['Name', 'Is Default', 'Source Self Mass?', 'Source Added Mass?',
                         'Source Load Patterns?', 'Load Pattern', 'Multiplier']

    # ('Name', 'Is Default', 'Include Lateral Mass?', 'Include Vertical Mass?', 'Lump Mass?', 'Source Self Mass?',
    #  'Source Added Mass?', 'Source Load Patterns?', 'Move Mass Centroid?', 'Move Ratio X', 'Move Ratio Y',
    #  'Load Pattern', 'Multiplier', 'GUID')

    TableData = ['mass', 'Yes', 'No', 'No', 'Yes', 'SELF WEIGHT', '1']

    [TableVersion, FieldsKeysIncluded, TableData,
     ret] = SapModel.DatabaseTables.SetTableForEditingArray(tTableKey, TableVersion,
                                                            fieldKeysIncluded, NumberRecords, TableData)

    [NumFatalErrors, NumErrorMsgs, NumWarnMsgs, NumInfoMsgs,
     ImportLog, ret] = SapModel.DatabaseTables.ApplyEditedTables(FillImportLog=True)
    return ret


# ======================================================================================================================


def setbucklingLoadCase(SapModel):
    [_, loadName, ret] = SapModel.LoadPatterns.GetNameList()

    grvLdPtrn = []
    grvLd = [1, 2, 3, 4]
    for ld in loadName:
        [ldtyp, ret] = SapModel.LoadPatterns.GetLoadType(ld)
        if ldtyp in grvLd:
            grvLdPtrn.append(ld)

    TableVersion = 1
    NumberRecords = len(grvLdPtrn)
    tTableKey = "Load Case Definitions - Buckling"
    fieldKeysIncluded = ['Name', 'Load Type', 'Load Name', 'Load SF', 'Number Modes']

    TableData = []
    for i, grL in enumerate(grvLdPtrn):
        if i == 0:
            TableData = TableData + ['buckling', 'Load', grL, "1", "4"]
        else:
            TableData = TableData + ['buckling', 'Load', grL, "1", None]

    TableData = tuple(TableData)
    [TableVersion, FieldsKeysIncluded, TableData,
     ret] = SapModel.DatabaseTables.SetTableForEditingArray(tTableKey, TableVersion,
                                                            fieldKeysIncluded, NumberRecords, TableData)

    [NumFatalErrors, NumErrorMsgs, NumWarnMsgs, NumInfoMsgs,
     ImportLog, ret] = SapModel.DatabaseTables.ApplyEditedTables(FillImportLog=True)
    return ret


# ======================================================================================================================

def setLoadPtrnCaseCombosForOutput(SapModel, rmCase=(3, 10), rmPtrn=None, rmAllCombo=False):
    [_, ptrnName, ret] = SapModel.LoadPatterns.GetNameList()
    [_, caseName, ret] = SapModel.LoadCases.GetNameList()
    try:
        [_, comboName, ret] = SapModel.RespCombo.GetNameList()
    except:
        comboName = None

    keepCase = []
    keepPtrn = []

    if rmCase is not None:
        for ld in caseName:
            [ldtyp, sbtyp, ret] = SapModel.LoadCases.GetTypeOAPI(ld)
            if not ldtyp in rmCase:
                keepCase.append(ld)

    if rmPtrn is not None:
        for ld in ptrnName:
            [ldtyp, ret] = SapModel.LoadPatterns.GetLoadType(ld)
            if not ldtyp in rmPtrn:
                keepPtrn.append(ld)
    else:
        keepPtrn = ptrnName

    ret = SapModel.DatabaseTables.SetLoadPatternsSelectedForDisplay(keepPtrn)
    ret = SapModel.DatabaseTables.SetLoadCasesSelectedForDisplay(keepCase)
    if not comboName is None:
        if rmAllCombo:
            ret = SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay([])
        else:
            ret = SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay(comboName)

    return ret


# ===============================================================================================================

def edit_diaphragm(SapModel):
    [_, diaphName, ret] = SapModel.Diaphragm.GetNameList()

    for d in diaphName:
        ret = SapModel.Diaphragm.SetDiaphragm(d, SemiRigid=False)
    return ret


# ===============================================================================================================

def setLoadCasetoRun(SapModel, notRunCase=[6]):
    # [_, ptrnName, ret] = SapModel.LoadPatterns.GetNameList()
    [_, caseName, ret] = SapModel.LoadCases.GetNameList()
    # try:
    #     [_, comboName, ret] = SapModel.RespCombo.GetNameList()
    # except:
    #     comboName = None
    notrunCase = notRunCase
    notRunLoad = []

    ret = SapModel.Analyze.SetRunCaseFlag(Name="", Run=True, All=True)

    for ld in caseName:
        [_, _, dtype, _, _, ret] = SapModel.LoadCases.GetTypeOAPI_1(ld)
        if dtype in notrunCase:
            notRunLoad.append(ld)
            ret = SapModel.Analyze.SetRunCaseFlag(ld, Run=False, All=False)

    return ret, notRunLoad
